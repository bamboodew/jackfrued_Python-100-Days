import math

from tempfile import NamedTemporaryFile

from openpyxl import Workbook, load_workbook

wb = Workbook()
ws = wb.active
ws1 = wb.create_sheet('MySheet')  # insert at the end (default)
ws2 = wb.create_sheet('MySheet1', 0)  # insert at first position
ws3 = wb.create_sheet('MySheet2', -1)  # insert at the penultimate position
ws.title = 'New Title'
ws.sheet_properties.tabColor = '1072BA'
ws3 = wb['New Title']
# print(wb.sheetnames)
# for sheet in wb:
#     print(sheet.title)
source = wb.active
target = wb.copy_worksheet(source)
"""
Accessing one cell
Now we know how to get a worksheet,
we can start modifying cells content.
Cells can be accessed directly as keys of the worksheet:
"""
ws['A4'] = 4
c = ws['A4']
# c.value = 'hello, world'
d = ws.cell(row=4, column=2, value=10)
# print(c)
# print(c.value)

# print(d)
# print(d.value)

for x in range(1, 101):
    for y in range(1, 101):
        ws.cell(row=x, column=y)

cell_range = ws['A1':'C2']  # 切片
# print(cell_range)  # 二维元组：先行再列
colC = ws['C']  # 共计1列、100行
# print(colC)  # 一维元组
col_range = ws['C:D']  # 共计2列、100行
# print(col_range)  # 先列再行
row10 = ws[10]
# print(row10)
row_range = ws[5:10]  # 先行再列
# print(row_range)

'''先行后列'''
for row in ws.iter_rows(min_row=1, max_col=3, max_row=2):
    for cell in row:
        print(cell)

print()
'''先列后行'''
for col in ws.iter_cols(min_row=1, max_col=3, max_row=2):
    for cell in col:
        print(cell)
print()
# print(tuple(ws.rows))
# print(tuple(ws.columns))

"""
Values only
If you just want the values from a worksheet you can use the Worksheet.values property. 
This iterates over all the rows in a worksheet but returns just the cell values:
"""
# for row in ws.values:
#     for value in row:
#         print(value)
"""
Both Worksheet.iter_rows() and Worksheet.iter_cols() can take the values_only parameter to return just the cell’s value:
"""
for row in ws.iter_rows(min_row=1, max_col=3, max_row=2, values_only=True):
    print(row)
"""
Data storage
Once we have a Cell, we can assign it a value:
"""
c.value = 'hello, world'
print(c)
print(c.value)
d.value = math.pi
print(d)
print(d.value)
"""
Saving to a file
The simplest and safest way to save a workbook is by using the Workbook.save() method of the Workbook object:
"""
# wb.save('balances.xlsx')  # This operation will overwrite existing files without warning.

with NamedTemporaryFile(delete=False) as tmp:
    wb.save(tmp.name)
    tmp.seek(0)
    stream = tmp.read()  # https://stackoverflow.com/questions/23212435/permission-denied-to-write-to-my-temporary-file
    tmp.close()

wb1 = load_workbook('document.xlsx')
wb1.template = True
wb1.save('document_template.xltx')

wb2 = load_workbook('document_template.xltx')
wb2.template = False
wb2.save('document2.xlsx')

wb3 = load_workbook('balances.xlsx')
print(wb3.sheetnames)
