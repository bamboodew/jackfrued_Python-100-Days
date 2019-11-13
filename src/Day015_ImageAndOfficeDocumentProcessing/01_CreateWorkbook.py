from openpyxl import Workbook

'''
There is no need to create a file on the filesystem to get started with openpyxl. 
Just import the Workbook class and start work:
'''
wb = Workbook()
'''
A workbook is always created with at least one worksheet. 
You can get it by using the Workbook.active property:
'''
ws = wb.active
'''
This is set to 0 by default. Unless you modify its value, 
you will always get the first worksheet by using this method.
'''
'''
You can create new worksheets using the Workbook.create_sheet() method:
'''
ws1 = wb.create_sheet('MySheet')  # insert at the end (default)
ws2 = wb.create_sheet('MySheet1', 0)  # insert at first position
ws3 = wb.create_sheet('MySheet2', -1)  # insert at the penultimate position
'''
Sheets are given a name automatically when they are created. 
They are numbered in sequence (Sheet, Sheet1, Sheet2, …). 
You can change this name at any time with the Worksheet.title property:
'''
ws.title = 'New Title'
'''
The background color of the tab holding this title is white by default.
You can change this providing an RRGGBB color code to the Worksheet.sheet_properties.tabColor attribute:
'''
ws.sheet_properties.tabColor = '1072BA'
'''
Once you gave a worksheet a name, you can get it as a key of the workbook:
'''
ws3 = wb['New Title']
'''
You can review the names of all worksheets of the workbook with the Workbook.sheetname attribute
'''
print(wb.sheetnames)
'''
You can loop through worksheets
'''
for sheet in wb:
    print(sheet.title)
'''
You can create copies of worksheets within a single workbook:
Workbook.copy_worksheet() method:
'''
source = wb.active
target = wb.copy_worksheet(source)
