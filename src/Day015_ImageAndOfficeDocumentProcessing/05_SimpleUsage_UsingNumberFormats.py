import datetime

from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.utils import FORMULAE

wb = Workbook()
ws = wb.active
ws['A1'] = datetime.datetime(2010, 7, 21)
print(ws['A1'].number_format)  # 单元格内容的格式

ws["A2"] = "=SUM(1,1)"

print('HEX2DEC' in FORMULAE)  # 检查有无该函数

ws.merge_cells('A3:D3')  # 合并单元格
# ws.unmerge_cells('A3:D3')
ws.merge_cells(start_row=4, start_column=1, end_row=5, end_column=5)
# ws.unmerge_cells(start_row=4, start_column=1, end_row=5, end_column=5)

# Inserting an image
ws['A6'] = 'You should see logo below'
img = Image('logo.png')  # 插入图片
ws.add_image(img, 'A7')

ws.column_dimensions.group('G', 'J', hidden=True)  # 隐藏列
ws.row_dimensions.group(20, 30, hidden=True)  # 隐藏行

wb.save('formula.xlsx')
